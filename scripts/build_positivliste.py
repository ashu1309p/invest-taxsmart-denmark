#!/usr/bin/env python3
"""
Build data/positivliste.json from SKAT's official positivliste (ABIS) Excel file.

Runs ONLY inside the GitHub Action (build time) — never in the visitor's browser.
SKAT offers no live API and the file is CORS-blocked client-side, so we snapshot it here
and the static page loads only the finished JSON.

Guards (per the improvement brief):
  - If the download fails, the workbook looks wrong, or the new ISIN count drops by more than
    DROP_THRESHOLD vs the committed file, FAIL LOUDLY and do NOT overwrite the good JSON.

Set the real file URL via the POSITIVLISTE_URL env var (recommended) or edit DEFAULT_URL.
SKAT's file is "Liste over aktiebaserede investeringsselskaber" on
https://skat.dk/erhverv/ekapital/vaerdipapirer  — confirm the current direct .xlsx link there.

Deps (Action only):  pip install requests openpyxl
"""
import io
import json
import os
import re
import sys
import time
from datetime import datetime, timezone

import requests
from openpyxl import load_workbook

DEFAULT_URL = ""  # TODO: paste SKAT's current .xlsx direct link, or set POSITIVLISTE_URL in the Action.
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "positivliste.json")
DROP_THRESHOLD = 0.25          # refuse to overwrite if the new list is >25% smaller
ISIN_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}[0-9]$")

# SKAT (and many gov sites) 403 or hang on the bare python-requests user-agent, so send
# browser-like headers and retry transient failures before giving up.
HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"),
    "Accept": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
               "application/octet-stream,*/*"),
    "Accept-Language": "da,en;q=0.8",
}
RETRIES = 4
RETRY_BACKOFF = 5              # seconds, multiplied by attempt number

def fail(msg):
    print("ERROR:", msg, file=sys.stderr)
    sys.exit(1)

def download(url):
    """GET the file with browser headers, retrying transient errors with backoff."""
    last_err = None
    for attempt in range(1, RETRIES + 1):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=60)
            resp.raise_for_status()
            return resp
        except Exception as e:
            last_err = e
            print(f"AUDIT download attempt {attempt}/{RETRIES} failed: {e}", file=sys.stderr)
            if attempt < RETRIES:
                time.sleep(RETRY_BACKOFF * attempt)
    fail(f"Download failed after {RETRIES} attempts: {last_err}")

def load_existing_count():
    try:
        with open(OUT, encoding="utf-8") as f:
            return int(json.load(f).get("count", 0))
    except Exception:
        return 0

def pick_sheet(wb):
    """Prefer a sheet whose title contains the current or next year, else the last sheet."""
    years = [str(datetime.now().year), str(datetime.now().year + 1)]
    for y in years:
        for name in wb.sheetnames:
            if y in name:
                return wb[name]
    return wb[wb.sheetnames[-1]]

def find_columns(ws):
    """Locate the ISIN column (and a name column if present) from the header row.

    ISIN match is a case-insensitive SUBSTRING so headers like "ISIN-kode" or
    "Fondskode (ISIN)" are found, not just an exact "isin". Returns the matched
    header texts too, so the run is auditable.
    """
    isin_col = name_col = None
    isin_hdr = name_hdr = ""
    for row in ws.iter_rows(min_row=1, max_row=8):
        for cell in row:
            raw = str(cell.value or "").strip()
            v = raw.lower()
            if isin_col is None and "isin" in v:
                isin_col, isin_hdr = cell.column, raw
            if name_col is None and ("navn" in v or "name" in v or "fond" in v):
                name_col, name_hdr = cell.column, raw
        if isin_col:
            return isin_col, name_col, row[0].row, isin_hdr, name_hdr
    return None, None, None, "", ""

def main():
    url = os.environ.get("POSITIVLISTE_URL", DEFAULT_URL).strip()
    if not url:
        fail("No POSITIVLISTE_URL set and DEFAULT_URL is empty. Provide SKAT's .xlsx link.")

    print(f"AUDIT source URL: {url}")
    resp = download(url)
    ctype = resp.headers.get("Content-Type", "?")
    print(f"AUDIT HTTP {resp.status_code}, {len(resp.content)} bytes, content-type {ctype!r}")

    # A 200 that returns HTML (login/error page) instead of the xlsx is the other common
    # silent failure — surface it clearly rather than letting openpyxl throw a vague error.
    head = resp.content[:4]
    if head[:2] != b"PK":
        fail(f"Response is not an .xlsx (zip) file — first bytes {head!r}, content-type {ctype!r}. "
             "The POSITIVLISTE_URL link likely changed or returned an error page.")

    try:
        wb = load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    except Exception as e:
        fail(f"Could not parse workbook (truncated or not an xlsx?): {e}")

    ws = pick_sheet(wb)
    print(f"AUDIT sheets available: {wb.sheetnames}")
    print(f"AUDIT chosen sheet: {ws.title!r}")

    isin_col, name_col, header_row, isin_hdr, name_hdr = find_columns(ws)
    if not isin_col:
        fail("Could not find an 'ISIN' column header in the sheet. Layout may have changed.")
    print(f"AUDIT header row: {header_row}")
    print(f"AUDIT ISIN column: {isin_col} (header {isin_hdr!r})")
    print(f"AUDIT name column: {name_col} (header {name_hdr!r})")

    isins = {}
    for row in ws.iter_rows(min_row=header_row + 1):
        cell = row[isin_col - 1]
        code = str(cell.value or "").strip().upper().replace(" ", "")
        if not ISIN_RE.match(code):
            continue
        name = ""
        if name_col and len(row) >= name_col:
            name = str(row[name_col - 1].value or "").strip()
        # Membership in SKAT's file == ON the positivliste.
        isins[code] = {"name": name, "status": "on"}

    count = len(isins)
    print(f"AUDIT extracted ISIN count: {count}")
    if count == 0:
        fail("Extracted 0 ISINs — refusing to overwrite the committed snapshot.")

    prev = load_existing_count()
    if prev and count < prev * (1 - DROP_THRESHOLD):
        fail(f"New count {count} is >{int(DROP_THRESHOLD*100)}% below previous {prev}. "
             "Suspected bad/truncated fetch — keeping the existing good JSON.")

    payload = {
        "listDate": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "fetchedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source": "skat.dk positivliste (ABIS)",
        "sample": False,
        "count": count,
        "isins": isins,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=0, sort_keys=True)
    print(f"Wrote {count} ISINs to {OUT} (prev {prev}).")

if __name__ == "__main__":
    main()
